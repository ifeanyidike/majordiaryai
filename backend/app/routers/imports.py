"""Bulk cow master-data import (.xlsx / .csv).

The real Excel column format is NOT confirmed yet, so this pipeline is built to
be tolerant: headers are normalized and matched against a flexible alias map,
dates/statuses are parsed leniently, and unrecognized columns are ignored (but
reported). When the real file arrives, the ONLY place that should need editing
is the COLUMN_ALIASES map below.
"""

import csv
import io
import uuid
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from fastapi import (
    APIRouter, Depends, File, HTTPException, Query, UploadFile, status,
)
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.auth import get_current_user, require_roles
from app.core.database import get_db
from app.models.models import Cow, CowStatus, Farm
from app.schemas.imports import ImportResult, ImportRowError
from app.services.access import check_farm_access, get_allowed_farm_ids

router = APIRouter()

# Max upload size (bytes). Anything larger is rejected with 413.
_MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB


# ---------------------------------------------------------------------------
# ADJUST THIS MAP WHEN THE REAL EXCEL FORMAT IS CONFIRMED.
# ---------------------------------------------------------------------------
# This is the single source of truth for turning spreadsheet column headers
# into Cow model fields. The keys are canonical Cow field names; each value is
# the list of accepted header spellings (already in NORMALIZED form: lowercase,
# trimmed, with runs of spaces / underscores / hyphens collapsed to a single
# underscore -- see _normalize_header). To support a new header spelling from
# the real file, just add it to the relevant list. To support a brand-new
# field, add a new entry here (and, if it needs special parsing, extend
# _extract_cow_fields below).
#
# Headers that match nothing here are ignored (and surfaced in
# ImportResult.ignored_columns) rather than causing an error.
# ---------------------------------------------------------------------------
COLUMN_ALIASES: Dict[str, List[str]] = {
    "ear_tag": ["ear_tag", "eartag", "tag", "cow_id", "id", "animal_id"],
    "breed": ["breed"],
    "date_of_birth": ["date_of_birth", "dob", "birth_date", "born"],
    "lactation_number": ["lactation_number", "lactation", "lact"],
    "status": ["status"],
    "current_program": ["current_program", "program"],
    "last_calving_date": ["last_calving_date", "calving_date", "fresh_date"],
    "last_insemination_date": [
        "last_insemination_date", "ai_date", "bred_date", "insemination_date",
    ],
    "due_date": ["due_date"],
    "dry_date": ["dry_date"],
    "notes": ["notes"],
}

# Optional per-row farm override column (matched by farm id or farm name).
# Kept separate from COLUMN_ALIASES because it does not map onto a plain Cow
# scalar field -- it selects which farm the row belongs to.
FARM_COLUMN_ALIASES: List[str] = ["farm", "farm_id", "farm_name", "farm_code"]

# Reverse lookup: normalized header -> canonical field. Built once at import.
_ALIAS_TO_FIELD: Dict[str, str] = {}
for _field, _aliases in COLUMN_ALIASES.items():
    for _alias in _aliases:
        _ALIAS_TO_FIELD[_alias] = _field

# Date fields that go through tolerant date parsing.
_DATE_FIELDS = frozenset([
    "date_of_birth", "last_calving_date", "last_insemination_date",
    "due_date", "dry_date",
])

# Accepted date string formats, tried in order.
#
# Slash and dash separated values are genuinely ambiguous: 03/04/2026 is
# March 4th to a US herd and April 3rd to a European one. Day-first used to be
# tried first, which silently moved every unambiguous US date by months --
# and these columns drive the +283 due date and the +223 dry-off, so a
# misread calving date puts a cow on the wrong worklist for a whole gestation.
# The client is a US operation (the farm timezone defaults to Eastern), so
# month-first wins the tie. A value that only parses day-first (13/04/2026)
# still falls through to the day-first pattern.
#
# The dotted form is European convention and stays day-first.
_DATE_FORMATS = [
    "%Y-%m-%d",
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d",
    "%m/%d/%Y",
    "%d/%m/%Y",
    "%m-%d-%Y",
    "%d-%m-%Y",
    "%d.%m.%Y",
    "%d %b %Y",
    "%d %B %Y",
]

# Spellings a herd manager actually types, mapped onto the enum. Anything
# outside this and the enum's own names is an error, not a silent default.
_STATUS_SYNONYMS = {
    "bred": CowStatus.inseminated,
    "ai": CowStatus.inseminated,
    "ai_ed": CowStatus.inseminated,
    "inseminated": CowStatus.inseminated,
    "served": CowStatus.inseminated,
    "preg": CowStatus.pregnant,
    "pregnant": CowStatus.pregnant,
    "in_calf": CowStatus.pregnant,
    "confirmed_pregnant": CowStatus.pregnant,
    "empty": CowStatus.open,
    "not_pregnant": CowStatus.open,
    "open": CowStatus.open,
    "dried_off": CowStatus.dry,
    "dry": CowStatus.dry,
    "milking": CowStatus.fresh,
    "fresh": CowStatus.fresh,
    "lactating": CowStatus.fresh,
    "heifer": CowStatus.heifer,
    "calf": CowStatus.calf,
    "on_program": CowStatus.needling,
    "on_protocol": CowStatus.needling,
    "needling": CowStatus.needling,
    "culled": CowStatus.cull,
    "cull": CowStatus.cull,
    "sold": CowStatus.sold,
    "dead": CowStatus.dead,
    "deceased": CowStatus.dead,
    "died": CowStatus.dead,
}


class RowError(ValueError):
    """A row is unusable and must be reported, not quietly defaulted.

    Every one of these was previously a silent fallback -- an unreadable date
    became None, an unknown status became `open`, an unmatched farm name became
    the default farm. Each corrupted a cow's record while the import reported
    success, which is worse than refusing the row: nobody goes looking for a
    problem the importer said it did not have.
    """


# ---------------------------------------------------------------------------
# Pure parsing helpers (no DB access -- unit-testable offline).
# ---------------------------------------------------------------------------
def _normalize_header(header) -> str:
    """Lowercase, strip, and collapse any run of spaces/underscores/hyphens
    into a single underscore. Returns '' for None/blank."""
    if header is None:
        return ""
    text = str(header).strip().lower()
    out = []
    prev_sep = False
    for ch in text:
        if ch in (" ", "\t", "_", "-"):
            if not prev_sep:
                out.append("_")
                prev_sep = True
        else:
            out.append(ch)
            prev_sep = False
    return "".join(out).strip("_")


def _build_header_map(headers: List) -> Tuple[Dict[str, int], Optional[int], List[str]]:
    """Map spreadsheet headers to fields.

    Returns (field_to_index, farm_col_index, ignored_columns). First occurrence
    of a field wins if a spelling appears twice. Blank headers are skipped
    silently; unrecognized non-blank headers go into ignored_columns (original
    text preserved).
    """
    field_to_index: Dict[str, int] = {}
    farm_col_index: Optional[int] = None
    ignored: List[str] = []
    for idx, raw in enumerate(headers):
        norm = _normalize_header(raw)
        if not norm:
            continue
        field = _ALIAS_TO_FIELD.get(norm)
        if field is not None:
            if field not in field_to_index:
                field_to_index[field] = idx
            continue
        if norm in FARM_COLUMN_ALIASES:
            if farm_col_index is None:
                farm_col_index = idx
            continue
        ignored.append(str(raw).strip() if raw is not None else "")
    return field_to_index, farm_col_index, ignored


def _cell_str(value) -> Optional[str]:
    """Normalize a raw cell to a trimmed string, or None if blank."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return str(value).strip() or None


def _parse_date(value, field: str) -> Optional[date]:
    """Parse a cell into a date. Blank -> None; unreadable -> RowError.

    Accepts date/datetime objects (openpyxl gives these for real date cells),
    ISO strings, and common m/d/y or d/m/y spellings.

    An unreadable date used to come back as None. For `last_calving_date` that
    silently erased the basis for the due date, the dry-off date and the
    post-calving worklist, and the import still counted the row a success.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_str(value)
    if not text:
        return None
    # Trailing time on an ISO value like "2024-01-02T00:00:00"
    text = text.replace("T", " ").strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise RowError(
        f"{field}: could not read the date {text!r}. Use YYYY-MM-DD or MM/DD/YYYY."
    )


def _parse_status(value) -> CowStatus:
    """Map a cell to a CowStatus. Blank -> open; unrecognized -> RowError.

    An unknown word used to become `open`. A sheet that says "Pregnant 60d"
    or "SOLD" therefore imported the cow as open to breed -- putting a sold
    animal, or one carrying a calf, onto the technician's breeding list.
    """
    text = _cell_str(value)
    if not text:
        return CowStatus.open
    key = text.strip().lower().replace("-", "_").replace(" ", "_")
    for member in CowStatus:
        if member.value == key or member.name == key:
            return member
    synonym = _STATUS_SYNONYMS.get(key)
    if synonym is not None:
        return synonym
    raise RowError(
        f"status: {text!r} is not a status we recognize. "
        f"Use one of: {', '.join(m.value for m in CowStatus)}."
    )


def _parse_lactation(value) -> int:
    """Parse lactation number as a non-negative int. Blank/invalid -> 0,
    negatives clamped to 0."""
    text = _cell_str(value)
    if not text:
        return 0
    try:
        n = int(float(text))
    except (ValueError, TypeError):
        return 0
    return n if n > 0 else 0


def _extract_cow_fields(raw: Dict[str, object]) -> Dict[str, object]:
    """Turn a {field: raw_cell} dict (already keyed by canonical field) into
    parsed Cow column values. `ear_tag` is assumed validated by the caller."""
    fields: Dict[str, object] = {}

    ear_tag = _cell_str(raw.get("ear_tag"))
    if ear_tag is not None:
        fields["ear_tag"] = ear_tag

    for name in ("breed", "current_program", "notes"):
        if name in raw:
            fields[name] = _cell_str(raw.get(name))

    for name in _DATE_FIELDS:
        if name in raw:
            fields[name] = _parse_date(raw.get(name), name)

    if "lactation_number" in raw:
        fields["lactation_number"] = _parse_lactation(raw.get("lactation_number"))

    if "status" in raw:
        fields["status"] = _parse_status(raw.get("status"))

    return fields


# ---------------------------------------------------------------------------
# File reading (header + rows) for .xlsx / .csv.
# ---------------------------------------------------------------------------
def _read_table(filename: str, content: bytes) -> Tuple[List, List[List]]:
    """Return (headers, data_rows) from an .xlsx or .csv byte payload.

    Raises HTTPException(422) for an unsupported extension.
    """
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        return _read_xlsx(content)
    if lower.endswith(".csv"):
        return _read_csv(content)
    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail="Unsupported file type; expected .xlsx or .csv",
    )


def _read_xlsx(content: bytes) -> Tuple[List, List[List]]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Could not read .xlsx file",
        )
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return [], []
    headers = list(rows[0])
    data = [list(r) for r in rows[1:]]
    return headers, data


def _read_csv(content: bytes) -> Tuple[List, List[List]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    headers = rows[0]
    data = [list(r) for r in rows[1:]]
    return headers, data


def _row_is_empty(row: List) -> bool:
    return all(_cell_str(c) is None for c in row)


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------
@router.post("/cows", response_model=ImportResult)
async def import_cows(
    farm_id: uuid.UUID = Query(..., description="Default farm for imported cows"),
    file: UploadFile = File(...),
    current_user: dict = Depends(require_roles("admin", "technician")),
    db: AsyncSession = Depends(get_db),
):
    """Bulk-upsert cows from an uploaded .xlsx or .csv file.

    Every row defaults to the query `farm_id`. If the sheet carries a
    recognized farm column and the value matches a farm the caller can access,
    that row is assigned to that farm instead. Upsert key is (farm_id, ear_tag).
    """
    if not await check_farm_access(db, current_user, farm_id):
        raise HTTPException(status_code=404, detail="Farm not found")

    filename = file.filename or "upload"
    lower = filename.lower()
    if not (lower.endswith(".xlsx") or lower.endswith(".csv")):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Unsupported file type; expected .xlsx or .csv",
        )

    content = await file.read()
    if len(content) > _MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="File too large (max 5 MB)",
        )

    headers, data_rows = _read_table(filename, content)
    field_to_index, farm_col_index, ignored_columns = _build_header_map(headers)

    # Resolve which farms the caller may target, for per-row farm override.
    farm_by_id, farm_by_name = await _accessible_farm_lookup(db, current_user)

    created = 0
    updated = 0
    skipped = 0
    total_rows = 0
    errors: List[ImportRowError] = []

    for offset, row in enumerate(data_rows):
        # 1-based row number including the header row.
        row_number = offset + 2
        if _row_is_empty(row):
            continue
        total_rows += 1

        # Build {canonical_field: raw_cell} for this row.
        raw: Dict[str, object] = {}
        for field, idx in field_to_index.items():
            raw[field] = row[idx] if idx < len(row) else None

        ear_tag = _cell_str(raw.get("ear_tag"))
        if not ear_tag:
            skipped += 1
            errors.append(ImportRowError(
                row=row_number, ear_tag=None, message="Missing ear_tag",
            ))
            continue

        try:
            # Parse before touching the database: a bad cell is a row we never
            # want to write, and reporting it costs no round trip.
            target_farm_id = _resolve_row_farm(
                row, farm_col_index, farm_by_id, farm_by_name, farm_id,
            )
            fields = _extract_cow_fields(raw)

            # SAVEPOINT per row. A plain rollback() here discarded EVERY row
            # staged so far while the counters kept counting, so a 200-row file
            # with one bad row at 150 persisted only the rows after it and
            # still reported ~199 successes. begin_nested unwinds this row only.
            async with db.begin_nested():
                result = await db.execute(
                    select(Cow)
                    .where(Cow.farm_id == target_farm_id, Cow.ear_tag == ear_tag)
                    .with_for_update()
                )
                existing = result.scalar_one_or_none()
                if existing is not None:
                    for name, value in fields.items():
                        if name == "ear_tag":
                            continue
                        setattr(existing, name, value)
                    updated += 1
                else:
                    cow = Cow(farm_id=target_farm_id, **fields)
                    db.add(cow)
                    await db.flush()
                    created += 1
        except RowError as exc:
            # Already phrased for whoever is holding the spreadsheet.
            skipped += 1
            errors.append(ImportRowError(
                row=row_number, ear_tag=ear_tag, message=str(exc),
            ))
            continue
        except Exception as exc:  # noqa: BLE001 -- per-row isolation
            skipped += 1
            errors.append(ImportRowError(
                row=row_number, ear_tag=ear_tag,
                # Raw driver text can leak schema details; keep it short and
                # attributable without echoing the whole exception chain.
                message=str(exc).splitlines()[0][:200],
            ))
            continue

    await db.commit()

    return ImportResult(
        filename=filename,
        total_rows=total_rows,
        created=created,
        updated=updated,
        skipped=skipped,
        ignored_columns=ignored_columns,
        errors=errors,
    )


async def _accessible_farm_lookup(
    db: AsyncSession, current_user: dict,
) -> Tuple[Dict[uuid.UUID, Farm], Dict[str, uuid.UUID]]:
    """Build id->Farm and lower(name)->id maps for farms the caller can access.
    Used only to validate an optional per-row farm override column."""
    allowed = await get_allowed_farm_ids(db, current_user)
    stmt = select(Farm)
    if allowed is not None:
        if not allowed:
            return {}, {}
        stmt = stmt.where(Farm.id.in_(allowed))
    farms = (await db.execute(stmt)).scalars().all()
    by_id: Dict[uuid.UUID, Farm] = {f.id: f for f in farms}
    by_name: Dict[str, uuid.UUID] = {}
    for f in farms:
        if f.name:
            by_name[f.name.strip().lower()] = f.id
    return by_id, by_name


def _resolve_row_farm(
    row: List,
    farm_col_index: Optional[int],
    farm_by_id: Dict[uuid.UUID, Farm],
    farm_by_name: Dict[str, uuid.UUID],
    default_farm_id: uuid.UUID,
) -> uuid.UUID:
    """Return the farm id for a row: the override column value if it names an
    accessible farm (by id or name), otherwise the default query farm_id.

    A farm cell naming a farm we cannot match used to fall through to the
    default farm. A typo, or a sheet covering a farm the caller cannot access,
    therefore filed those cows onto whichever farm the upload was started from
    -- reported as imported, on a herd they do not belong to, and only
    discoverable by someone noticing extra ear tags. It is a RowError now.
    """
    if farm_col_index is None or farm_col_index >= len(row):
        return default_farm_id
    value = _cell_str(row[farm_col_index])
    if not value:
        return default_farm_id
    try:
        as_uuid = uuid.UUID(value)
        if as_uuid in farm_by_id:
            return as_uuid
    except (ValueError, AttributeError):
        pass
    by_name = farm_by_name.get(value.strip().lower())
    if by_name is not None:
        return by_name
    raise RowError(
        f"farm: {value!r} is not a farm you have access to. "
        "Leave the column blank to use the farm you are importing into."
    )


@router.get("/cows/template")
async def cows_template(
    current_user: dict = Depends(get_current_user),
):
    """Download a CSV template with canonical headers and one example row so
    users know the expected shape. Available to any authenticated user."""
    columns = list(COLUMN_ALIASES.keys())
    example = {
        "ear_tag": "A123",
        "breed": "Holstein",
        "date_of_birth": "2022-03-15",
        "lactation_number": "2",
        "status": "open",
        "current_program": "ovsynch",
        "last_calving_date": "2024-01-10",
        "last_insemination_date": "2024-02-01",
        "due_date": "2024-11-01",
        "dry_date": "2024-09-01",
        "notes": "Example row -- delete before importing",
    }
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    writer.writerow([example.get(c, "") for c in columns])
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=cow_import_template.csv"},
    )
